import json
import os
import re
import requests
import urllib.parse
import boto3
from PIL import Image
import pytesseract
from openpyxl import load_workbook
import csv


# Get environment vars
secret_sf_creds_name = os.getenv("SECRET_SF_CREDS_NAME")
base_url_sf = os.getenv("BASE_URL_SF")
file_download_path = os.getenv("FILE_DOWNLOAD_PATH")
bucket_name = os.getenv("BUCKET_NAME")
bucket_path_fw_ds = os.getenv("BUCKET_PATH_FW_DS")
secret_fw_creds_name = os.getenv("SECRET_FW_CREDS_NAME")
fw_chatflow = os.getenv("FW_CHATFLOW")
supported_formats = json.loads(os.getenv("SUPPORTED_FORMATS"))
supported_formats_img = json.loads(os.getenv("SUPPORTED_FORMATS_IMG"))
supported_formats_all = supported_formats + supported_formats_img


# Set Tesseract env
os.environ["PATH"] += os.pathsep + "/opt/bin"
os.environ["TESSDATA_PREFIX"] = "/opt/tessdata"


# Initialize the SecretsManager client
sm = boto3.client('secretsmanager')
# Initialize the S3 client
s3 = boto3.client('s3')



def sf_get_token():
    # Get SF credentials
    print("Getting SM creds...")
    res = sm.get_secret_value(SecretId=secret_sf_creds_name)

    client_id = json.loads(res["SecretString"])["consumer-key"]
    client_secret = json.loads(res["SecretString"])["consumer-secret"]

    # Prepare request parameters
    url_auth = base_url_sf + "/services/oauth2/token"

    payload_auth = {
        'grant_type': 'client_credentials',
        'client_id': client_id,
        'client_secret': client_secret
    }

    print("Getting SF token...")
    res = requests.post(url_auth, headers={"Content-Type":"application/x-www-form-urlencoded"}, data=payload_auth)

    if res.status_code != 200:
        raise Exception(f"Failed to retrieve token: {res.status_code} {res.reason}")
    else:
        print("token retrieval succeeded.")

    ### extracting token from auth call
    res_auth = json.loads(res.content)
    token = res_auth["access_token"]

    return token




def dl_sf_file(doc_id, token):
    url_download = base_url_sf + f"/{file_download_path}/{doc_id}/content"

    ### Call to Salesforce API
    print("Downloading file...")
    res = requests.head(url_download, headers = {"Content-Type": "application/json", "Authorization":"Bearer " + token})
    
    if res.status_code != 200:
        raise Exception(f"Failed to get headers for file: {res.status_code} {res.reason}")

    filename = re.search(r'filename="(.+)"', res.headers["Content-Disposition"]).group(1)
    filename_decoded = urllib.parse.unquote(filename)
    filename_base, filename_ext = os.path.splitext(filename_decoded)
    file_size = int(res.headers["Content-Length"])
    
    filetype = res.headers.get("Content-Type")
    print(f"Content-Type: {filetype}")

    if filename_ext not in supported_formats_all:
        raise Exception(f"Unsupported file extension: {filename_ext}")

    if file_size > 10 * 1024 * 1024:
        raise Exception(f"File exceeded 10MB size imit: {file_size / 1024 / 1024} MB")

    res = requests.get(url_download, headers = {"Content-Type": "application/json", "Authorization":"Bearer " + token})

    if res.status_code != 200:
        raise Exception(f"Failed to download file: {res.status_code} {res.reason}")
    
    print(f"Saving file: {filename_decoded} as generic /tmp/download...")
    with open(f"/tmp/download", "wb") as file:
        # Use iter_content to write the file in chunks
        for chunk in res.iter_content(chunk_size=8192):
            if chunk:  # filter out keep-alive new chunks
                file.write(chunk)
    
    return filename



def upload_files_s3(rec_id, doc_id, filename):
    filename_decoded = urllib.parse.unquote(filename)
    filename_base, filename_ext = os.path.splitext(filename_decoded)

    # Upload the file - flowise (general)
    print("Uploading original file...")
    if not filename_base.startswith("sffile_"):
        filename_base = "sffile_" + filename_base
    try:
        file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}{filename_ext}"
        s3.upload_file("/tmp/download", bucket_name, file_path)
        print(f"File {file_path} uploaded to {bucket_name}")
    except Exception as e:
        print(f"Found error while uploading {file_path}: {e}")

    # Upload the file - flowise (image extracted text)
    if filename_ext in supported_formats_img:
        print("Uploading image file's extracted text for upsertion...")
        with open("/tmp/image.txt", 'w') as f:
            txt_file = pytesseract.image_to_string(Image.open("/tmp/download"))
            f.write(txt_file)
        try:
            file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}.txt".replace("sffile", "sfimg")
            s3.upload_file("/tmp/image.txt", bucket_name, file_path)
            print(f"File {file_path} uploaded to {bucket_name}")
        except Exception as e:
            print(f"Found error while uploading {file_path}: {e}")
    elif filename_ext == ".xlsx":
        print("Uploading xslx file's extracted csv for upsertion...")
        os.rename("/tmp/download", "/tmp/download.xlsx") # openpyxl requires an extension
        wb = load_workbook("/tmp/download.xlsx")
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Open the output CSV file
            with open(f"/tmp/excel_{sheet_name}.csv", 'w', newline="") as f:
                writer = csv.writer(f)
                # Write the rows of the sheet
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)
            try:
                file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}_{sheet_name}.csv".replace("sffile", "sfxl")
                s3.upload_file(f"/tmp/excel_{sheet_name}.csv", bucket_name, file_path)
                print(f"File {file_path} uploaded to {bucket_name}")
            except Exception as e:
                print(f"Found error while uploading {file_path}: {e}")

    return file_path



def fw_get_api_key():
    # Get FW api key
    print("Getting SM creds...")
    res = sm.get_secret_value(SecretId=secret_fw_creds_name)

    return res["SecretString"]



def load_process_upsert(file_path, orig_filename, rec_id, fw_api_key):
    # UPSERT VECTOR DATA
    print("Upserting vector data...")
    if orig_filename.endswith(".xlsx"):
        # if it's an excel file, use a prefix that will match all the file's sheets
        file_path = "_".join(file_path.split("_")[:-1]) + "_"
    
    res = requests.post(
        f"https://d2br9m4wtztkg9.cloudfront.net/api/v1/vector/upsert/{fw_chatflow}",
        headers={"Authorization":f"Bearer {fw_api_key}","Content-Type":"application/json"},
        json={"overrideConfig":{"prefix":f"{file_path}","metadata":{"source": "/".join(file_path.split("/")[1:]), "record_id": rec_id}}}
    )

    if res.status_code != 200:
        raise Exception(f"Error in upsertion procedure, got from API: {res.status_code}: {res.reason}")
    else:
        print("Document Store vector data upsertion succeeded.")