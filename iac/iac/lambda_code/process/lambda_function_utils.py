import json
import os
import requests
import sys
import urllib.parse
import boto3
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract
from openpyxl import load_workbook
import csv
from fitz import open as fitz_open
from docx import Document



# Get environment vars
common_prefix = os.getenv("COMMON_PREFIX")
env = os.getenv("ENV")
bucket_name = os.getenv("BUCKET_NAME")
bucket_path_fw_ds = os.getenv("BUCKET_PATH_FW_DS")
secret_fw_creds_name = os.getenv("SECRET_FW_CREDS_NAME")
fw_chatflow = os.getenv("FW_CHATFLOW")
supported_formats = json.loads(os.getenv("SUPPORTED_FORMATS"))
supported_formats_img = json.loads(os.getenv("SUPPORTED_FORMATS_IMG"))


# Set Tesseract env
os.environ["PATH"] += os.pathsep + "/opt/bin"
os.environ["TESSDATA_PREFIX"] = "/opt/tessdata"


# Initialize the SecretsManager client
sm = boto3.client('secretsmanager')
# Initialize the S3 client
s3 = boto3.client('s3')
# Initialize the SSM client
ssm = boto3.client('ssm')



def create_text_file(text):
    print("Creating text file with the text included in the event...")
    with open("/tmp/event.txt", 'w') as f:
        f.write(text)



def extract_txt_from_pdf():
    print("Starting PDF text extraction...")
    input_path = "/tmp/download"
    output_path = "/tmp/extracted.txt"

    with fitz_open(input_path) as pdf_file:
        with open(output_path, 'w') as txt_file:
            for page in pdf_file:
                # Extract blocks of text
                blocks = page.get_text("blocks")
                # Sort blocks by their vertical position
                blocks.sort(key=lambda b: b[1])  # b[1] is the y-coordinate
                for block in blocks:
                    txt_file.write(block[4] + '\n')  # b[4] is the text content



def upload_files_s3(rec_id, filename, doc_id=None):
    if doc_id is None:
        print("Uploading generated event text file...")
        try:
            file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename}"
            s3.upload_file("/tmp/event.txt", bucket_name, file_path)
            print(f"File {file_path} uploaded to {bucket_name}")
        except Exception as e:
            print(f"Found error while uploading {file_path}: {e}")

    else:
        filename_decoded = urllib.parse.unquote(filename)
        filename_base, filename_ext = os.path.splitext(filename_decoded)

        # Upload the file - flowise (general)
        if not filename_base.startswith("sffile_"):
            filename_base = "sffile_" + filename_base
        
        try:
            print("Uploading original file...")
            file_path_full = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}{filename_ext}"
            s3.upload_file("/tmp/download", bucket_name, file_path_full)
            print(f"File {file_path_full} uploaded to {bucket_name}")
        except Exception as e:
            print(f"Found error while uploading {file_path_full}: {e}")

        # Upload the file - flowise (image extracted text)
        if filename_ext.lower() in supported_formats_img:
            print("Preparing image for text extraction...")
            image = Image.open("/tmp/download")
            image_format = image.format
            image = image.convert("L") # apply grayscale
            enhancer = ImageEnhance.Contrast(image)
            image = enhancer.enhance(2) # apply contrast
            image = image.filter(ImageFilter.SHARPEN) # sharpen
            image = image.resize((int(image.width * 1.5), int(image.height * 1.5))) # magnify
            image.save("/tmp/download_enhanced", format=image_format)

            print("Starting image text extraction...")
            with open("/tmp/extracted.txt", 'w') as f:
                txt_file = pytesseract.image_to_string(Image.open("/tmp/download_enhanced"))
                f.write(txt_file)

            try:
                print("Uploading image file's extracted text for upsertion...")
                file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}.txt".replace("sffile", "sfimg")
                s3.upload_file("/tmp/extracted.txt", bucket_name, file_path)
                print(f"File {file_path} uploaded to {bucket_name}")
            except Exception as e:
                print(f"Found error while uploading {file_path}: {e}")

        # Upload the file - flowise (xlsx extracted csvs)
        elif filename_ext.lower() == ".xlsx":
            print("Starting XLSX text extraction...")
            os.rename("/tmp/download", "/tmp/download.xlsx") # openpyxl requires an extension
            wb = load_workbook("/tmp/download.xlsx")

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # Open the output CSV file
                with open(f"/tmp/excel_{sheet_name}.csv", 'w', newline="") as f:
                    writer = csv.writer(f)
                    # Write the rows of the sheet
                    for row in sheet.iter_rows(values_only=True):
                        writer.writerow(row)

                try:
                    print("Uploading XLSX file's extracted csv sheet for upsertion...")
                    file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}_{sheet_name}.csv".replace("sffile", "sfxl")
                    s3.upload_file(f"/tmp/excel_{sheet_name}.csv", bucket_name, file_path)
                    print(f"File {file_path} uploaded to {bucket_name}")
                except Exception as e:
                    print(f"Found error while uploading {file_path}: {e}")
                    sys.exit(1)

        # Upload the file - flowise (pdf extracted text)
        elif filename_ext.lower() == ".pdf":
            extract_txt_from_pdf() # not using Flowise doc store extractor as it is quite faulty

            try:
                print("Uploading PDF file's extracted text for upsertion...")
                file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}.txt".replace("sffile", "sfpdf")
                s3.upload_file("/tmp/extracted.txt", bucket_name, file_path)
                print(f"File {file_path} uploaded to {bucket_name}")
            except Exception as e:
                print(f"Found error while uploading {file_path}: {e}")

        # Upload the file - flowise (docx extracted text)
        elif filename_ext.lower() == ".docx":
            print("Starting DOCX text extraction...")
            text = '\n'.join([paragraph.text for paragraph in Document("/tmp/download").paragraphs])
            with open("/tmp/extracted.txt", 'w', encoding='utf-8') as file:
                file.write(text)
            
            try:
                print("Uploading DOCX file's extracted text for upsertion...")
                file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}.txt".replace("sffile", "sfdocx")
                s3.upload_file("/tmp/extracted.txt", bucket_name, file_path)
                print(f"File {file_path} uploaded to {bucket_name}")
            except Exception as e:
                print(f"Found error while uploading {file_path}: {e}")

    return file_path, file_path_full



def fw_get_api_key():
    print("Getting FW api key from SM...")
    res = sm.get_secret_value(SecretId=secret_fw_creds_name)

    return res["SecretString"]



def upsert_vector(cf_distro_domain, chatflow, fw_api_key, file_path, file_path_source, rec_id):
    print("Upserting vector data...")
    try:
        res = requests.post(
            f"https://{cf_distro_domain}/api/v1/vector/upsert/{chatflow["id"]}",
            headers={"Authorization":f"Bearer {fw_api_key}","Content-Type":"application/json"},
            json={"overrideConfig":{"prefix":f"{file_path}","metadata":{"source": file_path_source, "record_id": rec_id}}}
        )
    except Exception as e:
        print(f"Found error while upserting: {e}")
        sys.exit(1)

    if res.status_code != 200:
        print(f"Error in upsertion procedure, got from API: {res.status_code}: {res.reason}")
        sys.exit(1)
    else:
        print("Document Store vector data upsertion succeeded.")
        #print(res.json())



def upsert_process(file_path, orig_filename, rec_id, fw_api_key):
    # First search for the target chatflow using Flowise API
    cf_distro_domain = ssm.get_parameter(Name=f"/{common_prefix}-{env}/pipeline/cf_distro_domain")['Parameter']['Value']

    try:
        res = requests.get(
            f"https://{cf_distro_domain}/api/v1/chatflows",
            headers={"Authorization":f"Bearer {fw_api_key}"}
        )
    except Exception as e:
        print(f"Found error while searching for chatflow: {e}")
        sys.exit(1)
    

    if res.status_code != 200:
        print(f"Error in chatflow search, got from API: {res.status_code}: {res.reason}")
        sys.exit(1)
    else:
        chatflow = next((d for d in res.json() if d["name"] == fw_chatflow), None)

        if chatflow != None:
            print(f"Found chatflow: {chatflow["id"]}")
        else:
            print(f"Could not find chatflow '{fw_chatflow}'")
            sys.exit(1)

    # if it's an excel file, use a slightly different approach for the prefix
    if orig_filename.endswith(".xlsx"):
        file_path = "_".join(file_path.split("_")[:-1]) # this is the prefix that matches all of the file's sheets
        file_path_source = file_path + ".xlsx"
    else:
        file_path_source = "/".join(file_path.split("/")[1:])
    
    # UPSERT VECTOR DATA
    upsert_vector(cf_distro_domain, chatflow, fw_api_key, file_path, file_path_source, rec_id)