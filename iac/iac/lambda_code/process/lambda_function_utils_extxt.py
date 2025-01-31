import json
import os
import sys
import boto3
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract
from openpyxl import load_workbook
import csv
from fitz import open as fitz_open
from docx import Document


# Set Tesseract env
os.environ["PATH"] += os.pathsep + "/opt/bin"
os.environ["TESSDATA_PREFIX"] = "/opt/tessdata"


# Get environment vars
env = os.getenv("ENV")
bucket_name = os.getenv("BUCKET_NAME")
bucket_path_fw_ds = os.getenv("BUCKET_PATH_FW_DS")
supported_formats = json.loads(os.getenv("SUPPORTED_FORMATS"))


# Set Tesseract env
os.environ["PATH"] += os.pathsep + "/opt/bin"
os.environ["TESSDATA_PREFIX"] = "/opt/tessdata"


# Initialize the SecretsManager client
sm = boto3.client('secretsmanager')
# Initialize the S3 client
s3 = boto3.client('s3')



def extract_txt_from_img(rec_id, filename_base, doc_id):
    print("Preparing image for text extraction...")
    image = Image.open("/tmp/download")
    image_format = image.format
    image = image.convert("L") # apply grayscale
    enhancer = ImageEnhance.Contrast(image)
    image = enhancer.enhance(2) # apply contrast
    image = image.filter(ImageFilter.SHARPEN) # sharpen
    image = image.resize((int(image.width * 1.5), int(image.height * 1.5))) # magnify
    image.save("/tmp/download_enhanced", format=image_format)

    print("Starting image text extraction...")
    with open("/tmp/extracted.txt", 'w') as f:
        txt_file = pytesseract.image_to_string(Image.open("/tmp/download_enhanced"))
        f.write(txt_file)

    try:
        print("Uploading image file's extracted text for upsertion...")
        file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}.txt".replace("sffile", "sfimg")
        s3.upload_file("/tmp/extracted.txt", bucket_name, file_path)
        print(f"File {file_path} uploaded to {bucket_name}")
    except Exception as e:
        print(f"Found error while uploading {file_path}: {e}")

    return file_path



def extract_txt_from_xlsx(rec_id, filename_base, doc_id):
    print("Starting XLSX text extraction...")
    os.rename("/tmp/download", "/tmp/download.xlsx") # openpyxl requires an extension
    wb = load_workbook("/tmp/download.xlsx")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Open the output CSV file
        with open(f"/tmp/excel_{sheet_name}.csv", 'w', newline="") as f:
            writer = csv.writer(f)
            # Write the rows of the sheet
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        try:
            print("Uploading XLSX file's extracted csv sheet for upsertion...")
            file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}_{sheet_name}.csv".replace("sffile", "sfxl")
            s3.upload_file(f"/tmp/excel_{sheet_name}.csv", bucket_name, file_path)
            print(f"File {file_path} uploaded to {bucket_name}")
        except Exception as e:
            print(f"Found error while uploading {file_path}: {e}")

    return file_path



def extract_txt_from_docx(rec_id, filename_base, doc_id):
    print("Starting DOCX text extraction...")
    text = '\n'.join([paragraph.text for paragraph in Document("/tmp/download").paragraphs])
    with open("/tmp/extracted.txt", 'w', encoding='utf-8') as file:
        file.write(text)
    
    try:
        print("Uploading DOCX file's extracted text for upsertion...")
        file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}.txt".replace("sffile", "sfdocx")
        s3.upload_file("/tmp/extracted.txt", bucket_name, file_path)
        print(f"File {file_path} uploaded to {bucket_name}")
    except Exception as e:
        print(f"Found error while uploading {file_path}: {e}")

    return file_path



def extract_txt_from_pdf(rec_id, filename_base, doc_id):
    print("Starting PDF text extraction...")
    input_path = "/tmp/download"
    output_path = "/tmp/extracted.txt"

    with fitz_open(input_path) as pdf_file:
        with open(output_path, 'w') as txt_file:
            for page in pdf_file:
                # Extract blocks of text
                blocks = page.get_text("blocks")
                # Sort blocks by their vertical position
                blocks.sort(key=lambda b: b[1])  # b[1] is the y-coordinate
                for block in blocks:
                    txt_file.write(block[4] + '\n')  # b[4] is the text content

    try:
        print("Uploading PDF file's extracted text for upsertion...")
        file_path = f"{bucket_path_fw_ds}/{rec_id}/{filename_base}_{doc_id}.txt".replace("sffile", "sfpdf")
        s3.upload_file("/tmp/extracted.txt", bucket_name, file_path)
        print(f"File {file_path} uploaded to {bucket_name}")
    except Exception as e:
        print(f"Found error while uploading {file_path}: {e}")

    return file_path